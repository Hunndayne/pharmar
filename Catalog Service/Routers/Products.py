from io import BytesIO
from typing import Annotated
from uuid import UUID

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlalchemy import func, select
from sqlalchemy.orm import selectinload

from Source.catalog import (
    apply_unit_role_invariants,
    ensure_unique_code,
    generate_next_code,
    get_primary_product_unit,
    get_drug_group_or_404,
    get_manufacturer_or_404,
    get_product_or_404,
    get_product_unit_or_404,
    normalize_unit_role,
    normalize_code,
    normalize_optional_string,
    paginate_scalars,
    parse_decimal,
    product_has_inventory_batches,
    product_has_inventory_history,
    product_has_sale_history,
    round_money_decimal,
    search_filter,
    to_product_detail,
    to_product_list_item,
    to_product_unit_response,
)
from Source.db.models import DrugGroup, Manufacturer, Product, ProductUnit
from Source.dependencies import (
    AccessToken,
    DbSession,
    ROLE_MANAGER,
    ROLE_OWNER,
    TokenUser,
    get_current_user,
    require_roles,
)
from Source.schemas.catalog import (
    BarcodeLookupResponse,
    PageResponse,
    ProductCreateRequest,
    ProductDetailResponse,
    ProductImportResult,
    ProductListItemResponse,
    ProductRoleUnitConfigRequest,
    ProductUnitRole,
    ProductUnitCreateRequest,
    ProductUnitResponse,
    ProductUnitUpdateRequest,
    ProductUpdateRequest,
)


router = APIRouter(prefix="/catalog", tags=["catalog-products"])

AnyUser = Annotated[TokenUser, Depends(get_current_user)]
ManagerOrOwner = Annotated[TokenUser, Depends(require_roles(ROLE_OWNER, ROLE_MANAGER))]
OwnerOnly = Annotated[TokenUser, Depends(require_roles(ROLE_OWNER))]


def _build_product_list_page(
    items: list[Product],
    total: int,
    page: int,
    size: int,
    pages: int,
) -> PageResponse[ProductListItemResponse]:
    return PageResponse[ProductListItemResponse](
        items=[to_product_list_item(item) for item in items],
        total=total,
        page=page,
        size=size,
        pages=pages,
    )


async def _validate_group_and_manufacturer(
    db: DbSession,
    group_id: UUID | None,
    manufacturer_id: UUID | None,
) -> tuple[DrugGroup | None, Manufacturer | None]:
    group = None
    manufacturer = None

    if group_id is not None:
        group = await get_drug_group_or_404(group_id, db)
    if manufacturer_id is not None:
        manufacturer = await get_manufacturer_or_404(manufacturer_id, db)
    return group, manufacturer


ROLE_DISPLAY_LABELS: dict[ProductUnitRole, str] = {
    "import": "import",
    "intermediate": "intermediate",
    "retail": "retail",
}


def _normalize_unit_name_or_400(value: str, field_label: str = "Unit name") -> str:
    normalized = normalize_optional_string(value)
    if normalized is None:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"{field_label} is required")
    return normalized


def _normalize_unit_role_or_400(
    value: str | None,
    *,
    fallback_is_base: bool = False,
    fallback_conversion: int | None = None,
) -> ProductUnitRole:
    normalized = normalize_unit_role(value)
    if normalized is not None:
        return normalized
    if fallback_is_base or fallback_conversion == 1:
        return "retail"
    return "import"


def _serialize_role_unit(
    *,
    role: ProductUnitRole,
    config: ProductRoleUnitConfigRequest,
    conversion_rate: int,
) -> dict[str, object]:
    if role == "retail" and not config.enabled:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Retail unit must stay enabled")
    return {
        "role": role,
        "unit_name": _normalize_unit_name_or_400(config.unit_name, f"{ROLE_DISPLAY_LABELS[role]} unit name"),
        "conversion_rate": conversion_rate,
        "barcode": normalize_optional_string(config.barcode),
        "selling_price": config.selling_price,
        "enabled": bool(config.enabled),
    }


def _build_role_unit_specs(payload: ProductCreateRequest | ProductUpdateRequest) -> list[dict[str, object]] | None:
    unit_config = payload.unit_config
    if unit_config is None:
        return None

    specs: list[dict[str, object]] = [
        _serialize_role_unit(role="retail", config=unit_config.retail, conversion_rate=1)
    ]
    import_config = unit_config.import_role
    intermediate_config = unit_config.intermediate
    import_enabled = bool(import_config and import_config.enabled)
    intermediate_enabled = bool(intermediate_config and intermediate_config.enabled)

    if intermediate_enabled and not import_enabled:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Intermediate unit requires the import unit to be enabled",
        )

    if intermediate_enabled and intermediate_config is not None:
        specs.append(
            _serialize_role_unit(
                role="intermediate",
                config=intermediate_config,
                conversion_rate=intermediate_config.conversion_to_lower_role,
            )
        )

    if import_enabled and import_config is not None:
        import_conversion = import_config.conversion_to_lower_role
        if intermediate_enabled and intermediate_config is not None:
            import_conversion *= intermediate_config.conversion_to_lower_role
        specs.append(
            _serialize_role_unit(
                role="import",
                config=import_config,
                conversion_rate=import_conversion,
            )
        )

    return specs


def _build_fallback_retail_spec(payload: ProductCreateRequest) -> list[dict[str, object]]:
    base_payload = payload.base_unit
    base_unit_name = _normalize_unit_name_or_400(base_payload.unit_name, "Retail unit name") if base_payload else "Viên"
    base_unit_price = base_payload.selling_price if base_payload else 0
    return [
        {
            "role": "retail",
            "unit_name": base_unit_name,
            "conversion_rate": 1,
            "barcode": normalize_optional_string(payload.barcode),
            "selling_price": base_unit_price,
            "enabled": True,
        }
    ]


def _sync_product_units_by_role(product: Product, specs: list[dict[str, object]]) -> None:
    desired_roles = {str(spec["role"]) for spec in specs}
    existing_by_role = {
        role: next((unit for unit in product.units if normalize_unit_role(unit.unit_role) == role), None)
        for role in ("retail", "intermediate", "import")
    }

    for spec in specs:
        role = str(spec["role"])
        unit = existing_by_role.get(role)
        if unit is None:
            unit = ProductUnit(product_id=product.id)
            product.units.append(unit)
            existing_by_role[role] = unit
        unit.unit_role = role
        unit.unit_name = str(spec["unit_name"])
        unit.conversion_rate = int(spec["conversion_rate"])
        unit.barcode = spec["barcode"] if isinstance(spec["barcode"], str) or spec["barcode"] is None else None
        unit.selling_price = spec["selling_price"]  # type: ignore[assignment]
        unit.is_active = bool(spec["enabled"])
        apply_unit_role_invariants(unit)

    for unit in product.units:
        role = normalize_unit_role(unit.unit_role)
        if role is None:
            if unit.is_active:
                unit.is_active = False
            continue
        if role not in desired_roles:
            unit.is_active = False
            apply_unit_role_invariants(unit)


@router.get("/products", response_model=PageResponse[ProductListItemResponse])
async def list_products(
    _: AnyUser,
    db: DbSession,
    search: str | None = Query(default=None),
    group_id: UUID | None = Query(default=None),
    manufacturer_id: UUID | None = Query(default=None),
    is_active: bool | None = Query(default=None),
    page: int = Query(default=1, ge=1),
    size: int = Query(default=20, ge=1, le=200),
) -> PageResponse[ProductListItemResponse]:
    stmt = (
        select(Product)
        .options(
            selectinload(Product.group),
            selectinload(Product.manufacturer),
            selectinload(Product.units),
        )
        .order_by(Product.created_at.desc())
    )
    stmt = search_filter(
        stmt,
        search,
        Product.code,
        Product.name,
        Product.barcode,
        Product.registration_number,
    )
    if group_id is not None:
        stmt = stmt.where(Product.group_id == group_id)
    if manufacturer_id is not None:
        stmt = stmt.where(Product.manufacturer_id == manufacturer_id)
    if is_active is not None:
        stmt = stmt.where(Product.is_active == is_active)

    rows, meta = await paginate_scalars(db, stmt, page, size)
    return _build_product_list_page(rows, meta.total, meta.page, meta.size, meta.pages)


@router.get("/products/search", response_model=list[ProductListItemResponse])
async def quick_search_products(
    _: AnyUser,
    db: DbSession,
    q: str = Query(min_length=1),
    limit: int = Query(default=20, ge=1, le=100),
) -> list[ProductListItemResponse]:
    stmt = (
        select(Product)
        .options(
            selectinload(Product.group),
            selectinload(Product.manufacturer),
            selectinload(Product.units),
        )
        .where(Product.is_active.is_(True))
        .order_by(Product.updated_at.desc())
    )
    stmt = search_filter(
        stmt,
        q,
        Product.code,
        Product.name,
        Product.barcode,
        Product.registration_number,
    ).limit(limit)
    rows = list((await db.scalars(stmt)).all())
    return [to_product_list_item(item) for item in rows]


@router.get("/products/barcode/{barcode}", response_model=BarcodeLookupResponse)
async def get_product_by_barcode(barcode: str, _: AnyUser, db: DbSession) -> BarcodeLookupResponse:
    normalized = barcode.strip()
    if not normalized:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Barcode is required")

    unit_stmt = (
        select(ProductUnit)
        .where(
            ProductUnit.barcode == normalized,
            ProductUnit.is_active.is_(True),
        )
        .options(
            selectinload(ProductUnit.product).selectinload(Product.group),
            selectinload(ProductUnit.product).selectinload(Product.manufacturer),
            selectinload(ProductUnit.product).selectinload(Product.units),
        )
    )
    unit = await db.scalar(unit_stmt)
    if unit is not None and unit.product is not None and unit.product.is_active:
        return BarcodeLookupResponse(
            product=to_product_list_item(unit.product),
            unit=to_product_unit_response(unit),
        )

    product_stmt = (
        select(Product)
        .where(
            Product.barcode == normalized,
            Product.is_active.is_(True),
        )
        .options(
            selectinload(Product.group),
            selectinload(Product.manufacturer),
            selectinload(Product.units),
        )
    )
    product = await db.scalar(product_stmt)
    if product is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Barcode not found")

    active_unit = get_primary_product_unit(product)
    if active_unit is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="No active unit found for product",
        )

    return BarcodeLookupResponse(
        product=to_product_list_item(product),
        unit=to_product_unit_response(active_unit),
    )


@router.post("/products", response_model=ProductDetailResponse, status_code=status.HTTP_201_CREATED)
async def create_product(
    payload: ProductCreateRequest,
    _: ManagerOrOwner,
    db: DbSession,
) -> ProductDetailResponse:
    code = normalize_code(payload.code) or await generate_next_code(db, Product, "T")
    await ensure_unique_code(db, Product, code)
    await _validate_group_and_manufacturer(db, payload.group_id, payload.manufacturer_id)

    product = Product(
        code=code,
        barcode=normalize_optional_string(payload.barcode),
        name=payload.name.strip(),
        active_ingredient=normalize_optional_string(payload.active_ingredient),
        registration_number=normalize_optional_string(payload.registration_number),
        group_id=payload.group_id,
        manufacturer_id=payload.manufacturer_id,
        instructions=normalize_optional_string(payload.instructions),
        note=normalize_optional_string(payload.note),
        vat_rate=payload.vat_rate,
        other_tax_rate=payload.other_tax_rate,
        is_active=payload.is_active,
    )
    db.add(product)
    await db.flush()
    desired_specs = _build_role_unit_specs(payload) or _build_fallback_retail_spec(payload)
    _sync_product_units_by_role(product, desired_specs)
    await db.commit()

    created = await get_product_or_404(product.id, db)
    return to_product_detail(created)


@router.put("/products/{product_id:uuid}", response_model=ProductDetailResponse)
async def update_product(
    product_id: UUID,
    payload: ProductUpdateRequest,
    _: ManagerOrOwner,
    db: DbSession,
) -> ProductDetailResponse:
    product = await get_product_or_404(product_id, db)
    updates = payload.model_dump(exclude_unset=True)

    if "code" in updates and payload.code is not None:
        code = normalize_code(payload.code)
        if code is None:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Code cannot be empty")
        await ensure_unique_code(db, Product, code, exclude_id=product.id)
        product.code = code

    if "group_id" in updates:
        if payload.group_id is not None:
            await get_drug_group_or_404(payload.group_id, db)
        product.group_id = payload.group_id

    if "manufacturer_id" in updates:
        if payload.manufacturer_id is not None:
            await get_manufacturer_or_404(payload.manufacturer_id, db)
        product.manufacturer_id = payload.manufacturer_id

    if "barcode" in updates:
        product.barcode = normalize_optional_string(payload.barcode)
    if "name" in updates and payload.name is not None:
        product.name = payload.name.strip()
    if "active_ingredient" in updates:
        product.active_ingredient = normalize_optional_string(payload.active_ingredient)
    if "registration_number" in updates:
        product.registration_number = normalize_optional_string(payload.registration_number)
    if "instructions" in updates:
        product.instructions = normalize_optional_string(payload.instructions)
    if "note" in updates:
        product.note = normalize_optional_string(payload.note)
    if "vat_rate" in updates and payload.vat_rate is not None:
        product.vat_rate = payload.vat_rate
    if "other_tax_rate" in updates and payload.other_tax_rate is not None:
        product.other_tax_rate = payload.other_tax_rate
    if "is_active" in updates and payload.is_active is not None:
        product.is_active = payload.is_active
    if "unit_config" in updates:
        desired_specs = _build_role_unit_specs(payload)
        if desired_specs is not None:
            _sync_product_units_by_role(product, desired_specs)

    await db.commit()
    updated = await get_product_or_404(product_id, db)
    return to_product_detail(updated)


@router.delete("/products/{product_id:uuid}")
async def delete_product(
    product_id: UUID,
    _: OwnerOnly,
    token: AccessToken,
    db: DbSession,
):
    product = await get_product_or_404(product_id, db)
    if await product_has_inventory_batches(product.id):
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="Cannot delete product because inventory still has stock for this product",
        )

    inventory_has_history = await product_has_inventory_history(product.id)
    sale_has_history = await product_has_sale_history(product.id, token)

    if not inventory_has_history and not sale_has_history:
        await db.delete(product)
        await db.commit()
        return {"message": "Product permanently deleted"}

    product.is_active = False
    for unit in product.units:
        unit.is_active = False

    await db.commit()
    return {"message": "Product deleted (soft delete)"}


@router.get("/products/{product_id:uuid}", response_model=ProductDetailResponse)
async def get_product(product_id: UUID, _: AnyUser, db: DbSession) -> ProductDetailResponse:
    item = await get_product_or_404(product_id, db)
    return to_product_detail(item)


@router.get("/products/{product_id:uuid}/units", response_model=list[ProductUnitResponse])
async def list_product_units(
    product_id: UUID,
    _: AnyUser,
    db: DbSession,
    include_inactive: bool = Query(default=False),
) -> list[ProductUnitResponse]:
    product = await get_product_or_404(product_id, db)
    units = sorted(
        product.units,
        key=lambda item: (
            0 if item.is_active else 1,
            {"import": 0, "intermediate": 1, "retail": 2}.get(normalize_unit_role(item.unit_role) or "", 99),
            item.conversion_rate,
            item.unit_name.casefold(),
        ),
    )
    if not include_inactive:
        units = [unit for unit in units if unit.is_active]
    return [to_product_unit_response(unit) for unit in units]


@router.post("/products/{product_id:uuid}/units", response_model=ProductUnitResponse, status_code=status.HTTP_201_CREATED)
async def create_product_unit(
    product_id: UUID,
    payload: ProductUnitCreateRequest,
    _: ManagerOrOwner,
    db: DbSession,
) -> ProductUnitResponse:
    product = await get_product_or_404(product_id, db)
    role = _normalize_unit_role_or_400(
        payload.unit_role,
        fallback_is_base=payload.is_base_unit,
        fallback_conversion=payload.conversion_rate,
    )
    existing_role = next(
        (
            unit
            for unit in product.units
            if unit.is_active and normalize_unit_role(unit.unit_role) == role
        ),
        None,
    )
    if existing_role is not None:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=f"Product already has an active {ROLE_DISPLAY_LABELS[role]} unit",
        )

    item = ProductUnit(
        product_id=product_id,
        unit_name=_normalize_unit_name_or_400(payload.unit_name),
        unit_role=role,
        conversion_rate=1 if role == "retail" else payload.conversion_rate,
        barcode=normalize_optional_string(payload.barcode),
        selling_price=payload.selling_price,
        is_active=payload.is_active,
    )
    apply_unit_role_invariants(item)
    db.add(item)
    await db.commit()
    await db.refresh(item)
    return to_product_unit_response(item)


@router.put("/products/{product_id:uuid}/units/{unit_id:uuid}", response_model=ProductUnitResponse)
async def update_product_unit(
    product_id: UUID,
    unit_id: UUID,
    payload: ProductUnitUpdateRequest,
    _: ManagerOrOwner,
    db: DbSession,
) -> ProductUnitResponse:
    await get_product_or_404(product_id, db)
    unit = await get_product_unit_or_404(product_id, unit_id, db)
    updates = payload.model_dump(exclude_unset=True)
    target_role = _normalize_unit_role_or_400(
        payload.unit_role if "unit_role" in updates else unit.unit_role,
        fallback_is_base=payload.is_base_unit if "is_base_unit" in updates and payload.is_base_unit is not None else unit.is_base_unit,
        fallback_conversion=payload.conversion_rate if "conversion_rate" in updates and payload.conversion_rate is not None else unit.conversion_rate,
    )
    if target_role != normalize_unit_role(unit.unit_role):
        conflicting_role = await db.scalar(
            select(ProductUnit.id).where(
                ProductUnit.product_id == product_id,
                ProductUnit.id != unit.id,
                ProductUnit.unit_role == target_role,
                ProductUnit.is_active.is_(True),
            )
        )
        if conflicting_role is not None:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail=f"Product already has an active {ROLE_DISPLAY_LABELS[target_role]} unit",
            )

    if "unit_name" in updates and payload.unit_name is not None:
        unit.unit_name = _normalize_unit_name_or_400(payload.unit_name)
    unit.unit_role = target_role
    if "conversion_rate" in updates and payload.conversion_rate is not None:
        unit.conversion_rate = payload.conversion_rate
    if "barcode" in updates:
        unit.barcode = normalize_optional_string(payload.barcode)
    if "selling_price" in updates and payload.selling_price is not None:
        unit.selling_price = payload.selling_price
    if "is_active" in updates and payload.is_active is not None:
        if target_role == "retail" and not payload.is_active:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Retail unit cannot be deactivated",
            )
        unit.is_active = payload.is_active
    apply_unit_role_invariants(unit)

    await db.commit()
    await db.refresh(unit)
    return to_product_unit_response(unit)


@router.delete("/products/{product_id:uuid}/units/{unit_id:uuid}")
async def delete_product_unit(
    product_id: UUID,
    unit_id: UUID,
    _: ManagerOrOwner,
    db: DbSession,
):
    await get_product_or_404(product_id, db)
    unit = await get_product_unit_or_404(product_id, unit_id, db)

    if normalize_unit_role(unit.unit_role) == "retail" or unit.is_base_unit:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Cannot delete retail unit",
        )

    unit.is_active = False
    await db.commit()
    return {"message": "Product unit deleted (soft delete)"}


@router.get("/units/barcode/{barcode}", response_model=BarcodeLookupResponse)
async def get_unit_by_barcode(barcode: str, _: AnyUser, db: DbSession) -> BarcodeLookupResponse:
    normalized = barcode.strip()
    if not normalized:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Barcode is required")

    stmt = (
        select(ProductUnit)
        .where(
            ProductUnit.barcode == normalized,
            ProductUnit.is_active.is_(True),
        )
        .options(
            selectinload(ProductUnit.product).selectinload(Product.group),
            selectinload(ProductUnit.product).selectinload(Product.manufacturer),
            selectinload(ProductUnit.product).selectinload(Product.units),
        )
    )
    unit = await db.scalar(stmt)
    if unit is None or unit.product is None or not unit.product.is_active:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Unit barcode not found")

    return BarcodeLookupResponse(
        product=to_product_list_item(unit.product),
        unit=to_product_unit_response(unit),
    )


_XLSX_MAGIC = b"PK\x03\x04"  # ZIP/XLSX magic bytes (first 4 bytes)
_MAX_IMPORT_FILE_SIZE = 10 * 1024 * 1024  # 10 MB


@router.post("/products/import", response_model=ProductImportResult)
async def import_products_from_excel(
    _: ManagerOrOwner,
    db: DbSession,
    file: UploadFile = File(...),
) -> ProductImportResult:
    filename = (file.filename or "").lower()
    if not filename.endswith(".xlsx"):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Only .xlsx files are supported")

    content = await file.read()

    if len(content) > _MAX_IMPORT_FILE_SIZE:
        raise HTTPException(status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE, detail="File too large (max 10 MB)")

    if len(content) < 4 or content[:4] != _XLSX_MAGIC:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid file format. The uploaded file is not a valid .xlsx workbook.",
        )

    workbook = load_workbook(filename=BytesIO(content), data_only=True)
    sheet = workbook.active

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Excel file is empty")

    header = [str(cell).strip().lower() if cell is not None else "" for cell in rows[0]]
    required = {"name", "base_unit_name", "base_unit_price"}
    if not required.issubset(set(header)):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Header must include: name, base_unit_name, base_unit_price",
        )

    index_map = {name: idx for idx, name in enumerate(header)}
    group_rows = list((await db.scalars(select(DrugGroup))).all())
    manufacturer_rows = list((await db.scalars(select(Manufacturer))).all())
    group_by_code = {item.code.upper(): item for item in group_rows}
    manufacturer_by_code = {item.code.upper(): item for item in manufacturer_rows}

    imported = 0
    errors: list[str] = []

    for row_index, row_values in enumerate(rows[1:], start=2):
        if row_values is None:
            continue

        def cell_value(key: str):
            idx = index_map.get(key)
            if idx is None or idx >= len(row_values):
                return None
            return row_values[idx]

        name = normalize_optional_string(str(cell_value("name")) if cell_value("name") is not None else None)
        if name is None:
            continue

        base_unit_name = normalize_optional_string(
            str(cell_value("base_unit_name")) if cell_value("base_unit_name") is not None else None
        )
        if base_unit_name is None:
            errors.append(f"Row {row_index}: base_unit_name is required")
            continue

        base_unit_price = round_money_decimal(cell_value("base_unit_price"))
        if base_unit_price < 0:
            errors.append(f"Row {row_index}: base_unit_price must be >= 0")
            continue
        vat_rate = parse_decimal(cell_value("vat_rate"))
        if vat_rate < 0 or vat_rate > 100:
            errors.append(f"Row {row_index}: vat_rate must be between 0 and 100")
            continue
        other_tax_rate = parse_decimal(cell_value("other_tax_rate"))
        if other_tax_rate < 0 or other_tax_rate > 100:
            errors.append(f"Row {row_index}: other_tax_rate must be between 0 and 100")
            continue

        code = normalize_code(str(cell_value("code")) if cell_value("code") is not None else None)
        if code is None:
            code = await generate_next_code(db, Product, "T")

        group_code = normalize_code(str(cell_value("group_code")) if cell_value("group_code") is not None else None)
        manufacturer_code = normalize_code(
            str(cell_value("manufacturer_code")) if cell_value("manufacturer_code") is not None else None
        )
        group_id = group_by_code[group_code].id if group_code in group_by_code else None
        manufacturer_id = manufacturer_by_code[manufacturer_code].id if manufacturer_code in manufacturer_by_code else None

        existing = await db.scalar(select(Product.id).where(Product.code == code))
        if existing is not None:
            errors.append(f"Row {row_index}: product code '{code}' already exists")
            continue

        try:
            product = Product(
                code=code,
                barcode=normalize_optional_string(
                    str(cell_value("barcode")) if cell_value("barcode") is not None else None
                ),
                name=name,
                active_ingredient=normalize_optional_string(
                    str(cell_value("active_ingredient")) if cell_value("active_ingredient") is not None else None
                ),
                registration_number=normalize_optional_string(
                    str(cell_value("registration_number")) if cell_value("registration_number") is not None else None
                ),
                group_id=group_id,
                manufacturer_id=manufacturer_id,
                instructions=normalize_optional_string(
                    str(cell_value("instructions")) if cell_value("instructions") is not None else None
                ),
                note=normalize_optional_string(str(cell_value("note")) if cell_value("note") is not None else None),
                vat_rate=vat_rate,
                other_tax_rate=other_tax_rate,
                is_active=True,
            )
            db.add(product)
            await db.flush()

            unit = ProductUnit(
                product_id=product.id,
                unit_name=base_unit_name,
                conversion_rate=1,
                barcode=normalize_optional_string(
                    str(cell_value("unit_barcode")) if cell_value("unit_barcode") is not None else None
                ),
                selling_price=base_unit_price,
                is_base_unit=True,
                is_active=True,
            )
            db.add(unit)
            await db.commit()
            imported += 1
        except Exception:
            await db.rollback()
            errors.append(f"Row {row_index}: failed to import product (internal error)")

    return ProductImportResult(
        imported=imported,
        failed=len(errors),
        errors=errors,
    )


@router.get("/products/export")
async def export_products(_: ManagerOrOwner, db: DbSession):
    stmt = (
        select(Product)
        .options(
            selectinload(Product.group),
            selectinload(Product.manufacturer),
            selectinload(Product.units),
        )
        .order_by(Product.code.asc())
    )
    products = list((await db.scalars(stmt)).all())

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Products"
    sheet.append(
        [
            "code",
            "barcode",
            "name",
            "active_ingredient",
            "registration_number",
            "group_code",
            "group_name",
            "manufacturer_code",
            "manufacturer_name",
            "instructions",
            "note",
            "vat_rate",
            "other_tax_rate",
            "base_unit_name",
            "base_unit_price",
            "is_active",
        ]
    )

    for product in products:
        base_unit = next((item for item in product.units if item.is_base_unit), None)
        sheet.append(
            [
                product.code,
                product.barcode,
                product.name,
                product.active_ingredient,
                product.registration_number,
                product.group.code if product.group else None,
                product.group.name if product.group else None,
                product.manufacturer.code if product.manufacturer else None,
                product.manufacturer.name if product.manufacturer else None,
                product.instructions,
                product.note,
                float(product.vat_rate),
                float(product.other_tax_rate),
                base_unit.unit_name if base_unit else None,
                float(base_unit.selling_price) if base_unit else None,
                product.is_active,
            ]
        )

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=catalog_products.xlsx"},
    )
