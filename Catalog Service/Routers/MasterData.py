from decimal import Decimal
from io import BytesIO
from typing import Annotated
from uuid import UUID

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlalchemy import func, select

from Source.catalog import (
    ensure_unique_code,
    generate_next_code,
    get_drug_group_or_404,
    get_manufacturer_or_404,
    get_supplier_or_404,
    normalize_code,
    normalize_optional_string,
    paginate_scalars,
    parse_decimal,
    round_money_decimal,
    search_filter,
    to_supplier_debt_history,
)
from Source.db.models import DrugGroup, Manufacturer, Product, Supplier, SupplierDebtHistory
from Source.dependencies import (
    DbSession,
    ROLE_MANAGER,
    ROLE_OWNER,
    TokenUser,
    get_current_user,
    require_roles,
)
from Source.schemas.catalog import (
    DrugGroupCreateRequest,
    DrugGroupResponse,
    DrugGroupUpdateRequest,
    ManufacturerCreateRequest,
    ManufacturerResponse,
    ManufacturerUpdateRequest,
    PageResponse,
    SupplierCreateRequest,
    SupplierDebtPaymentRequest,
    SupplierDebtResponse,
    SupplierDebtHistoryResponse,
    SupplierResponse,
    SupplierUpdateRequest,
    ProductImportResult,
)


router = APIRouter(prefix="/catalog", tags=["catalog-master-data"])

AnyUser = Annotated[TokenUser, Depends(get_current_user)]
ManagerOrOwner = Annotated[TokenUser, Depends(require_roles(ROLE_OWNER, ROLE_MANAGER))]
OwnerOnly = Annotated[TokenUser, Depends(require_roles(ROLE_OWNER))]

_XLSX_MAGIC = b"PK\x03\x04"
_MAX_IMPORT_FILE_SIZE = 10 * 1024 * 1024


def _parse_excel_bool(value: object, default: bool = True) -> bool:
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    if not text:
        return default
    return text in {"1", "true", "yes", "y", "active", "đang hoạt động", "hoat dong", "x"}


def _cell_value(row_values: tuple[object, ...], index_map: dict[str, int], key: str):
    idx = index_map.get(key)
    if idx is None or idx >= len(row_values):
        return None
    return row_values[idx]


@router.get("/drug-groups", response_model=PageResponse[DrugGroupResponse])
async def list_drug_groups(
    _: AnyUser,
    db: DbSession,
    search: str | None = Query(default=None),
    is_active: bool | None = Query(default=None),
    page: int = Query(default=1, ge=1),
    size: int = Query(default=20, ge=1, le=200),
) -> PageResponse[DrugGroupResponse]:
    stmt = select(DrugGroup).order_by(DrugGroup.created_at.desc())
    stmt = search_filter(stmt, search, DrugGroup.code, DrugGroup.name)
    if is_active is not None:
        stmt = stmt.where(DrugGroup.is_active == is_active)

    rows, meta = await paginate_scalars(db, stmt, page, size)
    return PageResponse[DrugGroupResponse](
        items=[DrugGroupResponse.model_validate(item) for item in rows],
        total=meta.total,
        page=meta.page,
        size=meta.size,
        pages=meta.pages,
    )


@router.get("/drug-groups/{group_id}", response_model=DrugGroupResponse)
async def get_drug_group(group_id: UUID, _: AnyUser, db: DbSession) -> DrugGroupResponse:
    item = await get_drug_group_or_404(group_id, db)
    return DrugGroupResponse.model_validate(item)


@router.post("/drug-groups", response_model=DrugGroupResponse, status_code=status.HTTP_201_CREATED)
async def create_drug_group(
    payload: DrugGroupCreateRequest,
    _: ManagerOrOwner,
    db: DbSession,
) -> DrugGroupResponse:
    code = normalize_code(payload.code) or await generate_next_code(db, DrugGroup, "DG")
    await ensure_unique_code(db, DrugGroup, code)

    item = DrugGroup(
        code=code,
        name=payload.name.strip(),
        description=normalize_optional_string(payload.description),
        is_active=payload.is_active,
    )
    db.add(item)
    await db.commit()
    await db.refresh(item)
    return DrugGroupResponse.model_validate(item)


@router.put("/drug-groups/{group_id}", response_model=DrugGroupResponse)
async def update_drug_group(
    group_id: UUID,
    payload: DrugGroupUpdateRequest,
    _: ManagerOrOwner,
    db: DbSession,
) -> DrugGroupResponse:
    item = await get_drug_group_or_404(group_id, db)
    updates = payload.model_dump(exclude_unset=True)

    if "code" in updates and payload.code is not None:
        code = normalize_code(payload.code)
        if code is None:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Code cannot be empty")
        await ensure_unique_code(db, DrugGroup, code, exclude_id=item.id)
        item.code = code

    if "name" in updates and payload.name is not None:
        item.name = payload.name.strip()

    if "description" in updates:
        item.description = normalize_optional_string(payload.description)

    if "is_active" in updates and payload.is_active is not None:
        item.is_active = payload.is_active

    await db.commit()
    await db.refresh(item)
    return DrugGroupResponse.model_validate(item)


@router.delete("/drug-groups/{group_id}")
async def delete_drug_group(group_id: UUID, _: OwnerOnly, db: DbSession):
    item = await get_drug_group_or_404(group_id, db)
    linked_product = await db.scalar(
        select(Product.id).where(
            Product.group_id == item.id,
            Product.is_active.is_(True),
        )
    )
    if linked_product is not None:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="Cannot delete drug group because active products are using it",
        )

    item.is_active = False
    await db.commit()
    return {"message": "Drug group deleted (soft delete)"}


@router.get("/manufacturers", response_model=PageResponse[ManufacturerResponse])
async def list_manufacturers(
    _: AnyUser,
    db: DbSession,
    search: str | None = Query(default=None),
    is_active: bool | None = Query(default=None),
    page: int = Query(default=1, ge=1),
    size: int = Query(default=20, ge=1, le=200),
) -> PageResponse[ManufacturerResponse]:
    stmt = select(Manufacturer).order_by(Manufacturer.created_at.desc())
    stmt = search_filter(stmt, search, Manufacturer.code, Manufacturer.name)
    if is_active is not None:
        stmt = stmt.where(Manufacturer.is_active == is_active)

    rows, meta = await paginate_scalars(db, stmt, page, size)
    return PageResponse[ManufacturerResponse](
        items=[ManufacturerResponse.model_validate(item) for item in rows],
        total=meta.total,
        page=meta.page,
        size=meta.size,
        pages=meta.pages,
    )


@router.get("/manufacturers/{manufacturer_id:uuid}", response_model=ManufacturerResponse)
async def get_manufacturer(manufacturer_id: UUID, _: AnyUser, db: DbSession) -> ManufacturerResponse:
    item = await get_manufacturer_or_404(manufacturer_id, db)
    return ManufacturerResponse.model_validate(item)


@router.post("/manufacturers", response_model=ManufacturerResponse, status_code=status.HTTP_201_CREATED)
async def create_manufacturer(
    payload: ManufacturerCreateRequest,
    _: ManagerOrOwner,
    db: DbSession,
) -> ManufacturerResponse:
    code = normalize_code(payload.code) or await generate_next_code(db, Manufacturer, "MFG")
    await ensure_unique_code(db, Manufacturer, code)

    item = Manufacturer(
        code=code,
        name=payload.name.strip(),
        country=normalize_optional_string(payload.country),
        address=normalize_optional_string(payload.address),
        phone=normalize_optional_string(payload.phone),
        is_active=payload.is_active,
    )
    db.add(item)
    await db.commit()
    await db.refresh(item)
    return ManufacturerResponse.model_validate(item)


@router.put("/manufacturers/{manufacturer_id:uuid}", response_model=ManufacturerResponse)
async def update_manufacturer(
    manufacturer_id: UUID,
    payload: ManufacturerUpdateRequest,
    _: ManagerOrOwner,
    db: DbSession,
) -> ManufacturerResponse:
    item = await get_manufacturer_or_404(manufacturer_id, db)
    updates = payload.model_dump(exclude_unset=True)

    if "code" in updates and payload.code is not None:
        code = normalize_code(payload.code)
        if code is None:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Code cannot be empty")
        await ensure_unique_code(db, Manufacturer, code, exclude_id=item.id)
        item.code = code

    if "name" in updates and payload.name is not None:
        item.name = payload.name.strip()
    if "country" in updates:
        item.country = normalize_optional_string(payload.country)
    if "address" in updates:
        item.address = normalize_optional_string(payload.address)
    if "phone" in updates:
        item.phone = normalize_optional_string(payload.phone)
    if "is_active" in updates and payload.is_active is not None:
        item.is_active = payload.is_active

    await db.commit()
    await db.refresh(item)
    return ManufacturerResponse.model_validate(item)


@router.delete("/manufacturers/{manufacturer_id:uuid}")
async def delete_manufacturer(manufacturer_id: UUID, _: OwnerOnly, db: DbSession):
    item = await get_manufacturer_or_404(manufacturer_id, db)
    linked_product = await db.scalar(
        select(Product.id).where(
            Product.manufacturer_id == item.id,
            Product.is_active.is_(True),
        )
    )
    if linked_product is not None:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="Cannot delete manufacturer because active products are using it",
        )

    item.is_active = False
    await db.commit()
    return {"message": "Manufacturer deleted (soft delete)"}


@router.get("/manufacturers/export")
async def export_manufacturers(_: ManagerOrOwner, db: DbSession):
    rows = list((await db.scalars(select(Manufacturer).order_by(Manufacturer.code.asc()))).all())

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Manufacturers"
    sheet.append(["code", "name", "country", "address", "phone", "is_active"])
    for item in rows:
        sheet.append([item.code, item.name, item.country, item.address, item.phone, item.is_active])

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=catalog_manufacturers.xlsx"},
    )


@router.post("/manufacturers/import", response_model=ProductImportResult)
async def import_manufacturers(
    _: ManagerOrOwner,
    db: DbSession,
    file: UploadFile = File(...),
) -> ProductImportResult:
    filename = (file.filename or "").lower()
    if not filename.endswith(".xlsx"):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Only .xlsx files are supported")

    content = await file.read()
    if len(content) > _MAX_IMPORT_FILE_SIZE:
        raise HTTPException(status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE, detail="File too large (max 10 MB)")
    if len(content) < 4 or content[:4] != _XLSX_MAGIC:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid .xlsx file")

    workbook = load_workbook(filename=BytesIO(content), data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Excel file is empty")

    headers = [str(cell).strip().lower() if cell is not None else "" for cell in rows[0]]
    if "name" not in headers:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Header must include: name")
    index_map = {name: idx for idx, name in enumerate(headers)}

    imported = 0
    errors: list[str] = []

    for row_index, row_values in enumerate(rows[1:], start=2):
        if row_values is None:
            continue
        name_raw = _cell_value(row_values, index_map, "name")
        name = normalize_optional_string(str(name_raw) if name_raw is not None else None)
        if not name:
            continue

        code_raw = _cell_value(row_values, index_map, "code")
        code = normalize_code(str(code_raw) if code_raw is not None else None)
        country = normalize_optional_string(str(_cell_value(row_values, index_map, "country") or ""))
        address = normalize_optional_string(str(_cell_value(row_values, index_map, "address") or ""))
        phone = normalize_optional_string(str(_cell_value(row_values, index_map, "phone") or ""))
        is_active = _parse_excel_bool(_cell_value(row_values, index_map, "is_active"), default=True)

        try:
            existing = None
            if code:
                existing = await db.scalar(select(Manufacturer).where(Manufacturer.code == code))
            if existing is None:
                existing = await db.scalar(
                    select(Manufacturer).where(func.lower(Manufacturer.name) == name.lower())
                )

            if existing is not None:
                existing.name = name
                existing.country = country
                existing.address = address
                existing.phone = phone
                existing.is_active = is_active
                await db.commit()
                imported += 1
                continue

            final_code = code or await generate_next_code(db, Manufacturer, "MFG")
            await ensure_unique_code(db, Manufacturer, final_code)

            item = Manufacturer(
                code=final_code,
                name=name,
                country=country,
                address=address,
                phone=phone,
                is_active=is_active,
            )
            db.add(item)
            await db.commit()
            imported += 1
        except Exception:
            await db.rollback()
            errors.append(f"Row {row_index}: failed to import manufacturer")

    return ProductImportResult(imported=imported, failed=len(errors), errors=errors)


@router.get("/suppliers", response_model=PageResponse[SupplierResponse])
async def list_suppliers(
    _: AnyUser,
    db: DbSession,
    search: str | None = Query(default=None),
    is_active: bool | None = Query(default=None),
    page: int = Query(default=1, ge=1),
    size: int = Query(default=20, ge=1, le=200),
) -> PageResponse[SupplierResponse]:
    stmt = select(Supplier).order_by(Supplier.created_at.desc())
    stmt = search_filter(stmt, search, Supplier.code, Supplier.name, Supplier.phone)
    if is_active is not None:
        stmt = stmt.where(Supplier.is_active == is_active)

    rows, meta = await paginate_scalars(db, stmt, page, size)
    return PageResponse[SupplierResponse](
        items=[SupplierResponse.model_validate(item) for item in rows],
        total=meta.total,
        page=meta.page,
        size=meta.size,
        pages=meta.pages,
    )


@router.get("/suppliers/{supplier_id:uuid}", response_model=SupplierResponse)
async def get_supplier(supplier_id: UUID, _: AnyUser, db: DbSession) -> SupplierResponse:
    item = await get_supplier_or_404(supplier_id, db)
    return SupplierResponse.model_validate(item)


@router.post("/suppliers", response_model=SupplierResponse, status_code=status.HTTP_201_CREATED)
async def create_supplier(
    payload: SupplierCreateRequest,
    _: ManagerOrOwner,
    db: DbSession,
) -> SupplierResponse:
    code = normalize_code(payload.code) or await generate_next_code(db, Supplier, "SUP")
    await ensure_unique_code(db, Supplier, code)

    item = Supplier(
        code=code,
        name=payload.name.strip(),
        address=normalize_optional_string(payload.address),
        phone=payload.phone.strip(),
        email=normalize_optional_string(str(payload.email) if payload.email else None),
        tax_code=normalize_optional_string(payload.tax_code),
        contact_person=normalize_optional_string(payload.contact_person),
        current_debt=payload.current_debt,
        is_active=payload.is_active,
        note=normalize_optional_string(payload.note),
    )
    db.add(item)
    await db.commit()
    await db.refresh(item)
    return SupplierResponse.model_validate(item)


@router.put("/suppliers/{supplier_id:uuid}", response_model=SupplierResponse)
async def update_supplier(
    supplier_id: UUID,
    payload: SupplierUpdateRequest,
    _: ManagerOrOwner,
    db: DbSession,
) -> SupplierResponse:
    item = await get_supplier_or_404(supplier_id, db)
    updates = payload.model_dump(exclude_unset=True)

    if "code" in updates and payload.code is not None:
        code = normalize_code(payload.code)
        if code is None:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Code cannot be empty")
        await ensure_unique_code(db, Supplier, code, exclude_id=item.id)
        item.code = code

    if "name" in updates and payload.name is not None:
        item.name = payload.name.strip()
    if "address" in updates:
        item.address = normalize_optional_string(payload.address)
    if "phone" in updates and payload.phone is not None:
        item.phone = payload.phone.strip()
    if "email" in updates:
        item.email = normalize_optional_string(str(payload.email) if payload.email else None)
    if "tax_code" in updates:
        item.tax_code = normalize_optional_string(payload.tax_code)
    if "contact_person" in updates:
        item.contact_person = normalize_optional_string(payload.contact_person)
    if "is_active" in updates and payload.is_active is not None:
        item.is_active = payload.is_active
    if "note" in updates:
        item.note = normalize_optional_string(payload.note)

    await db.commit()
    await db.refresh(item)
    return SupplierResponse.model_validate(item)


@router.delete("/suppliers/{supplier_id:uuid}")
async def delete_supplier(supplier_id: UUID, _: OwnerOnly, db: DbSession):
    item = await get_supplier_or_404(supplier_id, db)
    linked_import = await db.scalar(
        select(SupplierDebtHistory.id).where(
            SupplierDebtHistory.supplier_id == item.id,
            SupplierDebtHistory.entry_type == "import",
        )
    )
    if linked_import is not None:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="Cannot delete supplier because import receipts are linked",
        )

    item.is_active = False
    await db.commit()
    return {"message": "Supplier deleted (soft delete)"}


@router.get("/suppliers/export")
async def export_suppliers(_: ManagerOrOwner, db: DbSession):
    rows = list((await db.scalars(select(Supplier).order_by(Supplier.code.asc()))).all())

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Suppliers"
    sheet.append(
        [
            "code",
            "name",
            "contact_person",
            "phone",
            "email",
            "tax_code",
            "address",
            "note",
            "current_debt",
            "is_active",
        ]
    )
    for item in rows:
        sheet.append(
            [
                item.code,
                item.name,
                item.contact_person,
                item.phone,
                item.email,
                item.tax_code,
                item.address,
                item.note,
                float(item.current_debt),
                item.is_active,
            ]
        )

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=catalog_suppliers.xlsx"},
    )


@router.post("/suppliers/import", response_model=ProductImportResult)
async def import_suppliers(
    _: ManagerOrOwner,
    db: DbSession,
    file: UploadFile = File(...),
) -> ProductImportResult:
    filename = (file.filename or "").lower()
    if not filename.endswith(".xlsx"):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Only .xlsx files are supported")

    content = await file.read()
    if len(content) > _MAX_IMPORT_FILE_SIZE:
        raise HTTPException(status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE, detail="File too large (max 10 MB)")
    if len(content) < 4 or content[:4] != _XLSX_MAGIC:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid .xlsx file")

    workbook = load_workbook(filename=BytesIO(content), data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Excel file is empty")

    headers = [str(cell).strip().lower() if cell is not None else "" for cell in rows[0]]
    required = {"name", "phone"}
    if not required.issubset(set(headers)):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Header must include: name, phone")
    index_map = {name: idx for idx, name in enumerate(headers)}

    imported = 0
    errors: list[str] = []

    for row_index, row_values in enumerate(rows[1:], start=2):
        if row_values is None:
            continue

        name_raw = _cell_value(row_values, index_map, "name")
        phone_raw = _cell_value(row_values, index_map, "phone")
        name = normalize_optional_string(str(name_raw) if name_raw is not None else None)
        phone = normalize_optional_string(str(phone_raw) if phone_raw is not None else None)
        if not name:
            continue
        if not phone:
            errors.append(f"Row {row_index}: phone is required")
            continue

        code_raw = _cell_value(row_values, index_map, "code")
        code = normalize_code(str(code_raw) if code_raw is not None else None)
        contact_person = normalize_optional_string(str(_cell_value(row_values, index_map, "contact_person") or ""))
        email = normalize_optional_string(str(_cell_value(row_values, index_map, "email") or ""))
        tax_code = normalize_optional_string(str(_cell_value(row_values, index_map, "tax_code") or ""))
        address = normalize_optional_string(str(_cell_value(row_values, index_map, "address") or ""))
        note = normalize_optional_string(str(_cell_value(row_values, index_map, "note") or ""))
        is_active = _parse_excel_bool(_cell_value(row_values, index_map, "is_active"), default=True)
        current_debt_raw = _cell_value(row_values, index_map, "current_debt")
        current_debt = round_money_decimal(current_debt_raw, Decimal("0"))
        if current_debt < 0:
            errors.append(f"Row {row_index}: current_debt must be >= 0")
            continue

        try:
            existing = None
            if code:
                existing = await db.scalar(select(Supplier).where(Supplier.code == code))
            if existing is None:
                existing = await db.scalar(
                    select(Supplier).where(
                        func.lower(Supplier.name) == name.lower(),
                        Supplier.phone == phone,
                    )
                )

            if existing is not None:
                existing.name = name
                existing.phone = phone
                existing.address = address
                existing.email = email
                existing.tax_code = tax_code
                existing.contact_person = contact_person
                existing.note = note
                existing.current_debt = current_debt
                existing.is_active = is_active
                await db.commit()
                imported += 1
                continue

            final_code = code or await generate_next_code(db, Supplier, "SUP")
            await ensure_unique_code(db, Supplier, final_code)
            item = Supplier(
                code=final_code,
                name=name,
                address=address,
                phone=phone,
                email=email,
                tax_code=tax_code,
                contact_person=contact_person,
                current_debt=current_debt,
                is_active=is_active,
                note=note,
            )
            db.add(item)
            await db.commit()
            imported += 1
        except Exception:
            await db.rollback()
            errors.append(f"Row {row_index}: failed to import supplier")

    return ProductImportResult(imported=imported, failed=len(errors), errors=errors)


@router.get("/suppliers/{supplier_id:uuid}/debt", response_model=SupplierDebtResponse)
async def get_supplier_debt(
    supplier_id: UUID,
    _: ManagerOrOwner,
    db: DbSession,
    page: int = Query(default=1, ge=1),
    size: int = Query(default=20, ge=1, le=200),
) -> SupplierDebtResponse:
    supplier = await get_supplier_or_404(supplier_id, db)

    stmt = (
        select(SupplierDebtHistory)
        .where(SupplierDebtHistory.supplier_id == supplier.id)
        .order_by(SupplierDebtHistory.created_at.desc())
    )
    rows, meta = await paginate_scalars(db, stmt, page, size)
    history = PageResponse[SupplierDebtHistoryResponse](
        items=[to_supplier_debt_history(item) for item in rows],
        total=meta.total,
        page=meta.page,
        size=meta.size,
        pages=meta.pages,
    )
    return SupplierDebtResponse(
        supplier_id=supplier.id,
        supplier_code=supplier.code,
        supplier_name=supplier.name,
        current_debt=supplier.current_debt,
        history=history,
    )


@router.post("/suppliers/{supplier_id:uuid}/debt/payment")
async def pay_supplier_debt(
    supplier_id: UUID,
    payload: SupplierDebtPaymentRequest,
    current_user: ManagerOrOwner,
    db: DbSession,
):
    supplier = await get_supplier_or_404(supplier_id, db)
    if payload.amount <= 0:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Payment amount must be positive")

    if Decimal(supplier.current_debt) < payload.amount:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Payment amount exceeds current debt",
        )

    new_balance = Decimal(supplier.current_debt) - payload.amount
    supplier.current_debt = new_balance

    history = SupplierDebtHistory(
        supplier_id=supplier.id,
        entry_type="payment",
        amount=payload.amount,
        balance_after=new_balance,
        reference_type="payment",
        reference_id=payload.reference_id,
        note=normalize_optional_string(payload.note),
        created_by=current_user.sub,
    )
    db.add(history)
    await db.commit()
    await db.refresh(history)
    await db.refresh(supplier)

    return {
        "message": "Supplier debt payment recorded",
        "supplier_id": str(supplier.id),
        "current_debt": supplier.current_debt,
        "entry": to_supplier_debt_history(history).model_dump(),
    }
